import os
from openpyxl import Workbook, load_workbook
from config.settings import FILTERED_URL_FILE
import openpyxl

def save_filtered_urls(filtered_urls, no_text_urls):
    """
    문자있음으로 필터링된 URL과 해당 판매자관리코드 및 필터링된 문자를 저장하는 함수.
    문자없음으로 필터링된 url과 해당 판매자관리코드 저장 

    중복된 URL은 저장하지 않으며
    '문자있음' 과 '문자없음' 탭으로 각각 저장
    
    Parameters:
    - filtered_urls: (판매자관리코드, URL, 필터링된 문자)로 이루어진 리스트
    - no_text_urls: "문자 없음"으로 판정된 URL 리스트
    """
    
    existing_urls = set()
    existing_no_text_urls = set()

    if os.path.exists(FILTERED_URL_FILE):
        workbook = load_workbook(FILTERED_URL_FILE)
        
        # "문자있음" 시트가 첫 번째 시트로 존재하는지 확인하고 불러오기
        if '문자있음' not in workbook.sheetnames:
            sheet = workbook.active
            sheet.title = "문자있음"  # 첫 번째 시트를 "문자있음"으로 이름 변경
        else:
            sheet = workbook["문자있음"]

        # 기존 필터링된 URL을 읽어 집합에 저장
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            existing_urls.add(row[1])

        # "문자없음" 시트가 없으면 생성, 있으면 불러오기
        if '문자없음' not in workbook.sheetnames:
            no_text_sheet = workbook.create_sheet("문자없음")
            no_text_sheet.append(["판매자관리코드", "URL"])  # 헤더 추가
        else:
            no_text_sheet = workbook["문자없음"]

        for row in no_text_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            existing_no_text_urls.add(row[1])

    else:
        # 파일이 없을 경우 새로 생성
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "문자있음"
        sheet.append(["판매자관리코드", "URL", "필터링된 문자"])
        no_text_sheet = workbook.create_sheet("문자없음")
        no_text_sheet.append(["판매자관리코드", "URL"])

    # 새로운 필터링된 URL 추가
    new_entries_text_count = 0
    new_entries_notext_count = 0
    for seller_code, url, detected_text in filtered_urls:
        if url not in existing_urls:
            sheet.append([seller_code, url, detected_text])
            new_entries_text_count += 1
            print(f"[디버그] 새 문자있음 URL 추가: 판매자관리코드: {seller_code}, URL: {url}, 필터링된 문자: {detected_text}")

    # 새로운 문자 없음 URL 추가
    for seller_code, url in no_text_urls:
        if url not in existing_no_text_urls:
            no_text_sheet.append([seller_code, url])
            new_entries_notext_count += 1
            print(f"[디버그] 새 문자없음 URL 추가: 판매자관리코드: {seller_code}, URL: {url} \n")

    # 변경사항 저장
    if new_entries_text_count > 0 or new_entries_notext_count > 0 :
        workbook.save(FILTERED_URL_FILE)
        print(f"[디버그]문자있음 URL {new_entries_text_count}개가 모음파일 글자있음탭에에 저장되었습니다.")
        print(f"[디버그]문자없음 URL {new_entries_notext_count}개가 모음파일 글자없음탭에에 저장되었습니다.")
    else:
        print(f"[디버그]새로운 필터링된 URL이 없어 변경사항이 저장되지 않았습니다.")
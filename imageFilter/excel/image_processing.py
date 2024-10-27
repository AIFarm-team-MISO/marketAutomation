import pandas as pd
from imageFilter.ocr_google.myImageFiltering import is_text_in_image
from utils.url_checker import log_processing

from config.settings import FILTERED_URL_FILE
from openpyxl import load_workbook
import openpyxl

import openpyxl
from openpyxl import Workbook

def load_filtered_urls(FILTERED_URL_FILE, url_list):
    """
    기존 필터링된 URL 목록과 주어진 URL 리스트를 비교하여 결과를 반환하는 함수.
    
    Parameters:
    - FILTERED_URL_FILE (str): 기존 필터링된 URL 목록이 저장된 파일 경로
    - url_list (list): 12열에 있는 이미지 URL 리스트
    
    Returns:
    - result (list): 
        ("중복-문자있음", image_url)        : 이전에 문자있어 필터링된 이미지 또는 
        ("중복-문자없음", image_url)   : 이전에 문자없음으로 필터링된 이미지 또는 
        ("새로운이미지", image_url)    : 기존에 필터링되지 않은 이미지 로 
        이루어진 리스트 
    """
    
    # 반환할 리스트 초기화
    result = []

    # 기존에 저장된 URL을 담을 집합
    existing_text_urls = set()  # 문자 있는 URL
    existing_no_text_urls = set()  # 문자 없는 URL

    try:
        # 기존 필터링된 URL 목록 파일을 열기
        workbook = openpyxl.load_workbook(FILTERED_URL_FILE)

        # "문자있음" 시트가 없으면 생성
        if '문자있음' not in workbook.sheetnames:
            text_sheet = workbook.create_sheet('문자있음')
            text_sheet.append(["판매자관리코드", "URL", "필터링된 문자"])  # 헤더 추가
        else:
            text_sheet = workbook['문자있음']

        # "문자없음" 시트가 없으면 생성
        if '문자없음' not in workbook.sheetnames:
            no_text_sheet = workbook.create_sheet('문자없음')
            no_text_sheet.append(["판매자관리코드", "URL"])  # 헤더 추가
        else:
            no_text_sheet = workbook['문자없음']

        # 문자 있는 URL 집합에 추가
        for row in text_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            url = row[1]
            if url:
                existing_text_urls.add(url)

        # 문자 없는 URL 집합에 추가
        for row in no_text_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            url = row[1]
            if url:
                existing_no_text_urls.add(url)

        print(f"기존 필터링된 URL {len(existing_text_urls) + len(existing_no_text_urls)}개가 로드되었습니다.")
    
    except FileNotFoundError:
        # 파일이 없으면 새로 생성
        print(f"필터링 모음파일이 존재하지 않으므로 새로 생성합니다: {FILTERED_URL_FILE}")
        workbook = Workbook()
        text_sheet = workbook.active
        text_sheet.title = '문자있음'
        text_sheet.append(["판매자관리코드", "URL", "필터링된 문자"])  # 헤더 추가
        no_text_sheet = workbook.create_sheet('문자없음')
        no_text_sheet.append(["판매자관리코드", "URL"])  # 헤더 추가

    # URL 리스트와 기존 목록 비교
    for image_url in url_list:
        if image_url in existing_text_urls:
            result.append(("중복-문자있음", image_url))
        elif image_url in existing_no_text_urls:
            result.append(("중복-문자없음", image_url))
        else:
            result.append(("새로운이미지", image_url))

    # 변경된 내용이 있으면 파일 저장
    workbook.save(FILTERED_URL_FILE)
    return result



def process_image_urls(sheet, image_column_index, seller_code_column_index, FILTERED_URL_FILE):
    """
    이미지 URL을 처리하여 문자 여부에 따라 필터링된 URL과 관련된 데이터를 분류하는 함수.
    
    Parameters:
    - sheet: 엑셀 시트 객체
    - image_column_index: 이미지 URL이 위치한 열(목록이미지) 의 인덱스
    - seller_code_column_index: 판매자관리코드가 위치한 열의 인덱스
    - FILTERED_URL_FILE: 기존 필터링된 URL 목록이 저장된 파일 경로
    
    Returns:
    - data: 필터링 결과를 포함한 리스트로 output 파일에 저장하기 위한 목록 
    - filtered_urls: 필터링된 URL 목록 (문자 포함)
    - no_text_urls: 문자 없는 URL 목록
    """
    
    # 반환할 리스트 초기화
    data = []  # 전체 처리 결과를 저장 (output 파일을 만들기 위해)
    filtered_urls = []  # 문자 포함 URL 목록 (이후 필터링url파일에 기록하기 위해)
    no_text_urls = []  # 문자 없음 URL 목록 (이후 필터링url파일에 기록하기 위해)

    # URL 리스트 생성 (3행부터 시작)
    url_list = [
        sheet.cell_value(idx, image_column_index) 
        for idx in range(2, sheet.nrows)  # 데이터가 3행부터 시작하므로 인덱스를 2부터 시작
        if pd.notna(sheet.cell_value(idx, image_column_index)) and 
           sheet.cell_value(idx, image_column_index).startswith('http')
    ]

    # 기존 필터링 목록과 비교하여 중복 여부를 확인
    # '목록이미지' 의 url들을 기존url모음 파일의 내용과 비교하여 url 대신에 내용을 추가
    # 이후 '목록이미지' 앞열에 필터링된 내용을 삽입 
    filtered_result = load_filtered_urls(FILTERED_URL_FILE, url_list)

    # 디버그: 필터링 결과 출력
    # print("\n[디버그] 기존파일과 url을 비교한 결과 목록:")
    # for status, image_url in filtered_result:
    #     print(f"상태: {status}, URL: {image_url}")

    # 필터링 결과를 하나씩 확인하여 상태에 따라 분류
    new_url_text_count = 0  # 새롭게 추가된 URL 중 문자있음 갯수
    new_url_notext_count = 0  # 새롭게 추가된 URL 중 문자없음 갯수

    for i, (status, image_url) in enumerate(filtered_result):
        idx = i + 2  # 데이터가 3행부터 시작하므로 인덱스 보정 필요
        seller_code = sheet.cell_value(idx, seller_code_column_index)  # 판매자 관리 코드 추출

        # 판매자관리코드가 비어 있을 경우 경고 메시지 출력
        if not seller_code:
            print(f"[경고] 판매자관리코드가 비어 있습니다. (URL: {image_url})")
            seller_code = "미확인"  # 판매자관리코드가 없을 경우 임시값으로 지정

        # URL 상태에 따라 분류 (이후 필터링url파일에 기록하기 위해)
        if status == "중복-문자있음":
            data.append((status, image_url))  # 처리 결과에 추가
            filtered_urls.append((seller_code, image_url, "중복-문자있음"))  # 문자 있음 목록에 추가
        elif status == "중복-문자없음":
            data.append((status, image_url))  # 처리 결과에 추가
            no_text_urls.append((seller_code, image_url))  # 문자 없음 목록에 추가
        else:
            # 새로운 URL인 경우 OCR을 사용하여 문자 필터링
            log_processing(idx, sheet.nrows, image_url)  # 현재 URL 처리 과정 로그
            detected_text = is_text_in_image(image_url)  # 이미지에서 텍스트 검출

            if detected_text:
                data.append(("중복-문자있음", image_url))  # 문자 있음 상태로 데이터에 추가
                filtered_urls.append((seller_code, image_url, detected_text))  # 문자 있음 목록에 추가
                new_url_text_count += 1  # 새롭게 추가된 URL 갯수 증가
            else:
                data.append(("중복-문자없음", image_url))  # 문자 없음 상태로 데이터에 추가
                no_text_urls.append((seller_code, image_url))  # 문자 없음 목록에 추가
                new_url_notext_count += 1  # 새롭게 추가된 URL 갯수 증가

    # 디버깅: 필터링된 URL 목록과 문자 없음 목록 출력
    print("\n[디버그] 문자 있음 필터링된 URL 목록 갯수 : " + str(len(filtered_urls)) + "개")
    # print("\n[디버그] 문자 있음 필터링된 URL 목록:")
    # for entry in filtered_urls:
    #     print(entry)

    print("[디버그] 문자 없음 필터링된 URL 목록 갯수 : " + str(len(no_text_urls)) +"개")
    # print("\n[디버그] 문자 없음 필터링된 URL 목록:")
    # for entry in no_text_urls:
    #     print(entry)

    # 디버그: 새롭게 추가된 URL 갯수 출력
    print(f"\n[디버그] 새롭게 추가된 URL 갯수: {new_url_text_count + new_url_notext_count}개")
    print(f"[디버그] 새롭게 추가된 URL 중 문자있음 갯수: {new_url_text_count}개")
    print(f"[디버그] 새롭게 추가된 URL 중 문자없음 갯수: {new_url_notext_count}개 \n")
    

    # 처리 결과 반환
    return data, filtered_urls, no_text_urls

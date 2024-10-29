from openpyxl import load_workbook
import win32com.client as win32
import psutil
import xlrd
from xlutils.copy import copy

import win32com.client as win32
from config.settings import YELLOW_COLOR_RGB  # 예시: (255, 255, 0)와 같은 RGB 값으로 설정 필요


import os
from config.settings import FILE_EXTENSION_xls

from openpyxl.styles import PatternFill
import win32com.client
import pythoncom

from openpyxl.utils import column_index_from_string, get_column_letter


def close_open_excel_files(file_name):
    """
    열려 있는 엑셀 파일을 찾아 강제로 닫는 함수.
    특정 파일 이름을 포함하는 엑셀 파일을 모두 닫아 파일 삭제가 가능하도록 처리합니다.

    Parameters:
    - file_name (str): 닫아야 할 엑셀 파일의 경로 또는 이름
    """
    # COM 라이브러리 사용을 위한 초기화 (멀티스레드 환경에서 안전하게 사용하기 위함)
    pythoncom.CoInitialize()
    
    # Excel Application 객체를 생성하여 엑셀 인스턴스를 가져옴
    xl = win32com.client.Dispatch("Excel.Application")
    
    # 열려 있는 모든 워크북 목록을 가져옴
    workbooks = xl.Workbooks

    try:
        # 열려 있는 모든 워크북을 반복하면서 지정된 파일이 열려 있는지 확인
        for wb in workbooks:
            # 파일 이름이 워크북의 전체 경로에 포함되어 있으면 닫기
            if file_name in wb.FullName:
                # 변경사항 저장 없이 워크북 닫기
                wb.Close(SaveChanges=False)
                print(f"엑셀 파일 닫힘: {wb.FullName}")

    except Exception as e:
        # 엑셀 파일 닫기 중 오류 발생 시 오류 메시지 출력
        print(f"엑셀 파일 닫기 중 오류 발생: {e}")

    finally:
        # Excel Application 종료 (모든 워크북 닫기 후 엑셀 프로그램 자체 종료)
        xl.Quit()
        
        # COM 객체 사용 종료 (자원 해제)
        pythoncom.CoUninitialize()

def process_all_excel_files(file_path):
    """
    폴더 내의 모든 엑셀 파일을 반복하여 파일 리스트를 반환하는 함수.
    기존의 .output 파일이 있으면 삭제한 후, 파일 경로와 이름을 리스트로 반환.
    
    Parameters:
    - file_path (str): 엑셀 파일이 위치한 폴더 경로
    
    Returns:
    - file_list (list): 파일 경로와 파일 이름 리스트 반환
    """
    file_list = []  # 처리할 파일 리스트를 저장할 리스트

    for file_name in os.listdir(file_path):
        # _output이 포함된 파일은 건너뜀 (이미 처리된 파일)
        if '_output' in file_name:
            continue
        
        # 엑셀 파일 확장자에 맞는 파일만 처리
        if file_name.endswith(FILE_EXTENSION_xls):
            base_file_name = os.path.splitext(file_name)[0]

            # output 파일 경로 설정
            output_file_path = os.path.join(file_path, f"{base_file_name}_output{FILE_EXTENSION_xls}")

            # 기존 output 파일이 있으면 삭제
            if os.path.exists(output_file_path):
                try:
                    os.remove(output_file_path)
                    print(f"기존 output 파일 삭제 완료: {output_file_path}")
                except PermissionError:
                    # 파일이 사용 중인 경우 엑셀 인스턴스를 닫고 다시 시도
                    print(f"파일이 열려있어 삭제할 수 없습니다. 엑셀 파일을 닫습니다: {output_file_path}")
                    close_open_excel_files(output_file_path)
                    try:
                        os.remove(output_file_path)
                        print(f"다시 시도 후 삭제 완료: {output_file_path}")
                    except Exception as e:
                        print(f"output 파일 삭제 중 오류 발생: {e}")
                        continue  # 삭제 실패 시 파일 리스트에 추가하지 않음

            # 파일 리스트에 추가
            file_list.append((file_path, base_file_name))

    return file_list

def column_letter_to_index(column_letter):
    """
    엑셀 열 문자 ('A', 'B', ..., 'AA', ...)를 0부터 시작하는 열 인덱스로 변환
    
    Parameters:
    - column_letter (str): 열 문자 (예: 'A', 'M', 'AA')

    Returns:
    - int: 0부터 시작하는 열 인덱스 (예: 'A' -> 0, 'M' -> 12)
    """
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char.upper()) - ord('A')) + 1
    return index - 1  # 0부터 시작하도록 -1

def insert_column_before(sheet, writable_sheet, column, new_column_title):
    """
    엑셀 시트에서 특정 열 앞에 새로운 열을 삽입하고, 첫 번째 행에 열 이름을 추가하는 함수.
    
    Parameters:
    - sheet: 읽기 전용 엑셀 시트 객체 (xlrd로 읽어들인 시트)
    - writable_sheet: 쓰기 가능한 엑셀 시트 객체 (xlutils.copy로 생성된 객체)
    - column (int 또는 str): 열 위치 (숫자 인덱스 또는 문자열 열 이름, 예: 'M' 또는 12)
    - new_column_title (str): 새 열의 제목 (예: "필터링결과")
    """
    # column이 문자열이면 숫자 인덱스로 변환
    if isinstance(column, str):
        column_index = column_letter_to_index(column)
    else:
        column_index = column

    # 모든 행에 대해, 기존 데이터를 오른쪽으로 이동하여 새로운 열 삽입
    for row_idx in range(1, sheet.nrows + 1):
        current_row_data = [sheet.cell_value(row_idx - 1, col_idx) for col_idx in range(sheet.ncols)]
        for col_idx in range(column_index, sheet.ncols):
            writable_sheet.write(row_idx - 1, col_idx + 1, current_row_data[col_idx])
    
    # 새로 삽입한 열에 대한 초기값 설정
    for row_idx in range(1, sheet.nrows + 1):
        writable_sheet.write(row_idx - 1, column_index, '')
    
    # 첫 번째 행(헤더)에 새로운 열 제목 추가
    writable_sheet.write(0, column_index, new_column_title)

import win32com.client as win32
import pythoncom

def apply_filter_and_sort_xls(output_file_path, sort_column, sort_direction='descending', sort_on='values'):
    """
    엑셀 파일을 열고 특정 열을 기준으로 필터링을 적용하고 정렬한 후, 특정 조건에 따라 색상을 적용하는 함수.
    
    Parameters:
    - output_file_path (str): 정렬할 엑셀 파일 경로
    - sort_column (str): 정렬할 열 (예: 'M'은 M열을 기준으로 정렬)
    - sort_direction (str): 정렬 방향 ('ascending' 또는 'descending', 기본값: 'descending')
    - sort_on (str): 정렬 기준 ('values' 또는 'color', 기본값: 'values')
    """
    
    # COM 라이브러리 사용을 위한 초기화 (특히 멀티스레드 환경에서 안정적인 사용을 위해 필요)
    pythoncom.CoInitialize()

    try:
        # 엑셀 애플리케이션 실행
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False  # 엑셀 창을 숨긴 상태로 실행

        # 엑셀 파일 열기
        workbook = excel.Workbooks.Open(output_file_path)
        sheet = workbook.Sheets(1)  # 첫 번째 시트 선택

        # 데이터가 포함된 마지막 열을 찾기
        last_column = sheet.UsedRange.Columns.Count  # 예: 데이터의 마지막 열이 'G'열이라면 last_column은 7이 됨
        
        # 필터 적용 (2행부터 마지막 열까지 적용)
        # A2 셀부터 마지막 열까지 범위를 지정하여 필터를 활성화
        sheet.Range(f"A2:{sheet.Cells(2, last_column).Address}").AutoFilter()

        # 엑셀 내부 상수 정의 (정렬 방향, 정렬 기준 등)
        xlSortOnValues = 0  # 값 기준 정렬
        xlSortOnCellColor = 1  # 색상 기준 정렬
        xlDescending = 2  # 내림차순 정렬 상수
        xlAscending = 1  # 오름차순 정렬 상수
        xlYes = 1  # 첫 번째 행을 헤더로 간주
        xlTopToBottom = 1  # 위에서 아래로 정렬

        # 정렬 방향 설정: 기본은 내림차순이며, sort_direction에 따라 방향 변경
        order = xlDescending if sort_direction == 'descending' else xlAscending

        # 정렬 기준 설정: 기본은 값 기준이며, sort_on에 따라 색상 기준 설정 가능
        sort_on_value = xlSortOnCellColor if sort_on == 'color' else xlSortOnValues

        # 정렬할 범위 설정 (A2부터 마지막 행까지)
        # A2 셀부터 마지막 데이터 행과 열까지 범위를 지정해 정렬 대상 영역 설정
        last_row = sheet.UsedRange.Rows.Count  # 데이터가 포함된 마지막 행 찾기
        sort_range = sheet.Range(f"A2:{sheet.Cells(last_row, last_column).Address}")

        # 정렬 필드를 초기화한 후, 새 필드를 추가하여 정렬을 수행할 열과 조건을 설정
        sheet.Sort.SortFields.Clear()
        sheet.Sort.SortFields.Add(
            Key=sheet.Range(f"{sort_column}2:{sort_column}{last_row}"),  # sort_column 기준으로 정렬
            SortOn=sort_on_value,  # 값 또는 색상 기준
            Order=order,  # 정렬 방향 (오름차순 또는 내림차순)
            DataOption=0  # 기본 정렬 옵션
        )

        # 정렬 설정 적용
        sheet.Sort.SetRange(sort_range)  # 정렬 범위 설정
        sheet.Sort.Header = xlYes  # 첫 번째 행을 헤더로 설정
        sheet.Sort.MatchCase = False  # 대소문자 구분 안 함
        sheet.Sort.Orientation = xlTopToBottom  # 위에서 아래로 정렬
        sheet.Sort.Apply()  # 정렬 적용

        # 정렬 완료 후 메시지 출력
        print(f"\n[디버그] 순환파일이 문자있음 우선으로 정렬 및 색상적용 완료\n")

        # 노란색을 적용할 Excel VBA 색상 코드 (여기서는 노란색 65535로 설정)
        yellow_excel_color = 65535

        # 색상 적용 루프: 조건에 맞는 셀에만 색상을 적용
        for row in range(2, last_row + 1):  # 2행부터 마지막 행까지 반복
            cell_value = sheet.Cells(row, 13).Value  # 13번째 열의 값을 가져와 조건 확인
            if cell_value == "중복-문자있음":  # 특정 조건 만족 시
                sheet.Rows(row).Interior.Color = yellow_excel_color  # 행 전체를 노란색으로 설정

        # 정렬 및 색상 적용이 완료된 파일을 저장
        workbook.Save()

    except Exception as e:
        # 오류 발생 시 메시지 출력
        print(f"파일 처리 중 오류 발생: {e}")

    finally:
        # 엑셀 파일을 닫고, 엑셀 프로세스 종료
        workbook.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()

        # 기타 자원 해제 작업 (필요에 따라 추가)
        clean_up_excel_process()


import xlwt

def apply_row_color_by_condition(writable_sheet, target_column, naming_list, color_name='yellow', condition_value=None):
    """
    xlwt 기반의 시트에서 리스트에 따라 특정 열의 값에 대해 행 색상을 적용하는 함수.
    
    Parameters:
    - writable_sheet: xlwt의 수정 가능한 sheet 객체
    - target_column (int): 색상 적용할 열 인덱스 (예: 4 -> E열)
    - naming_list (list): 행의 개수와 셀 값 참조용으로 사용
    - color_name (str): 색상 이름 (예: 'yellow', 'red' 등) 
    - condition_value (str, optional): 특정 문자열 값 (없을 경우 모든 값이 있는 행에 색상 적용)
    """
    # 색상 이름을 사용해 xlwt 스타일 생성
    pattern = xlwt.Pattern()
    pattern.pattern = xlwt.Pattern.SOLID_PATTERN
    pattern.pattern_fore_colour = xlwt.Style.colour_map.get(color_name, xlwt.Style.colour_map['yellow'])  # 기본값 노란색

    style = xlwt.XFStyle()
    style.pattern = pattern

    # 경고 메시지 출력 (모든 값이 있는 행에 색상을 적용할 때)
    if condition_value is None:
        print("[경고] condition_value=None이므로 해당 열의 모든 값이 있는 행에 색상을 적용합니다.")
    
    # naming_list의 길이를 사용하여 각 행에 조건에 따라 색상 적용
    for row, cell_value in enumerate(naming_list, start=2):  # 3행부터 시작
        if condition_value is None:
            # 모든 값이 있는 셀에 색상 적용
            if cell_value is not None and cell_value != "":
                writable_sheet.write(row, target_column, cell_value, style)
        else:
            # 특정 조건을 만족하는 셀에만 색상 적용
            if cell_value == condition_value:
                writable_sheet.write(row, target_column, cell_value, style)

    print("[디버그] 조건에 따라 색상 적용 완료")




def apply_filter_only_xls(output_file_path, sort_column, sort_direction='descending', sort_on='values'):
    """
    특정 열을 기준으로 필터링 및 정렬을 적용하는 함수 (색상 변경 없음).
    
    Parameters:
    - output_file_path (str): 정렬할 엑셀 파일 경로
    - sort_column (str): 정렬할 열 (예: 'M'은 M열을 기준으로 정렬)
    - sort_direction (str): 정렬 방향 ('ascending' 또는 'descending', 기본값: 'descending')
    - sort_on (str): 정렬 기준 ('values' 또는 'color', 기본값: 'values')
    """

    pythoncom.CoInitialize()
    try:
        # 엑셀 애플리케이션 실행
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False

        # 엑셀 파일 열기
        workbook = excel.Workbooks.Open(output_file_path)
        sheet = workbook.Sheets(1)

        # 필터 적용
        last_column = sheet.UsedRange.Columns.Count
        sheet.Range(f"A2:{sheet.Cells(2, last_column).Address}").AutoFilter()

        # 엑셀 상수 정의
        xlSortOnValues = 0
        xlSortOnCellColor = 1
        xlDescending = 2
        xlAscending = 1
        xlYes = 1
        xlTopToBottom = 1

        # 정렬 방향 설정
        order = xlDescending if sort_direction == 'descending' else xlAscending
        sort_on_value = xlSortOnCellColor if sort_on == 'color' else xlSortOnValues

        # 데이터 범위 설정
        last_row = sheet.UsedRange.Rows.Count
        sort_range = sheet.Range(f"A3:{sheet.Cells(last_row, last_column).Address}")

        # 정렬 필드 초기화 후 새 필드 추가
        sheet.Sort.SortFields.Clear()
        sheet.Sort.SortFields.Add(
            Key=sheet.Range(f"{sort_column}3:{sort_column}{last_row}"),
            SortOn=sort_on_value,
            Order=order,
            DataOption=0
        )

        # 정렬 수행
        sheet.Sort.SetRange(sort_range)
        sheet.Sort.Header = xlYes
        sheet.Sort.MatchCase = False
        sheet.Sort.Orientation = xlTopToBottom
        sheet.Sort.Apply()

        # 정렬된 파일을 저장
        workbook.Save()
        print("[디버그] 정렬 작업 완료")

    except Exception as e:
        print(f"정렬 작업 중 오류 발생: {e}")
    finally:
        workbook.Close(False)
        excel.Quit()
        pythoncom.CoUninitialize()









def clean_up_excel_process():
    """
    백그라운드에서 실행 중인 엑셀 프로세스를 종료하는 함수.
    
    작업 순서:
    1. 현재 실행 중인 프로세스 목록을 확인.
    2. 'excel.exe' 프로세스가 실행 중이면 이를 종료.
    """
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # 'EXCEL.EXE' 프로세스를 찾기
            if proc.info['name'].lower() == 'excel.exe':
                # 프로세스가 여전히 실행 중인지 확인
                if proc.is_running():
                    proc.kill()  # 프로세스 종료
                    print(f"[디버그] Excel 프로세스 종료됨 (pid={proc.info['pid']})")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            # 프로세스가 이미 종료되었거나 접근 권한이 없는 경우 예외 처리
            continue


# RGB 값을 Excel 색상 코드로 변환하는 함수 추가
def rgb_to_excel_color(rgb):
    return rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
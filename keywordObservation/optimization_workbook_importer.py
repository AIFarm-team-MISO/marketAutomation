from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from keywordObservation.tag_text_utils import (
    normalize_keyword,
    parse_tag_cell,
    stable_fingerprint,
)


class OptimizationWorkbookImportError(RuntimeError):
    pass


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "seller_product_code": (
        "판매자 상품코드",
        "상품코드",
        "판매자상품코드",
    ),
    "sale_code": (
        "판매코드",
        "상품번호",
    ),
    "keyword": (
        "메인키워드",
        "키워드",
    ),
    "original_product_name": (
        "원본상품명",
        "기본상품명",
        "원본 상품명",
    ),
    "processed_product_name": (
        "가공상품명",
        "가공 상품명",
    ),
    "tag_cell": (
        "검색태그",
        "검색 태그",
        "태그",
    ),
    "category": (
        "카테고리",
        "카테고리명",
    ),
}


REQUIRED_FIELDS = {
    "keyword",
    "original_product_name",
    "processed_product_name",
    "tag_cell",
}



def discover_optimization_workbooks(
    input_dir: Path,
    *,
    extensions: Iterable[str] = (".xlsx", ".xlsm", ".xls"),
) -> list[Path]:
    normalized_extensions = {
        str(extension).lower()
        if str(extension).startswith(".")
        else "." + str(extension).lower()
        for extension in extensions
    }

    if not input_dir.exists():
        return []

    return sorted(
        path
        for path in input_dir.iterdir()
        if path.is_file()
        and path.suffix.lower() in normalized_extensions
        and not path.name.startswith("~$")
    )



def _normalize_header(value: Any) -> str:
    return normalize_keyword(value).replace(" ", "")



def _build_header_map(
    headers: list[Any],
) -> dict[str, int]:
    normalized_headers = [
        _normalize_header(value)
        for value in headers
    ]

    header_map: dict[str, int] = {}

    for field, aliases in HEADER_ALIASES.items():
        normalized_aliases = {
            _normalize_header(alias)
            for alias in aliases
        }

        for index, header in enumerate(normalized_headers):
            if header in normalized_aliases:
                header_map[field] = index
                break

    # 일부 시트는 첫 번째 열의 헤더가 비어 있지만 판매자 상품코드가 들어 있다.
    if "seller_product_code" not in header_map and headers:
        if not normalize_keyword(headers[0]):
            header_map["seller_product_code"] = 0

    return header_map



def _cell_value(
    row: list[Any],
    header_map: dict[str, int],
    field: str,
) -> Any:
    index = header_map.get(field)

    if index is None or index >= len(row):
        return None

    return row[index]



def _build_record(
    *,
    source_file: Path,
    sheet_name: str,
    row_number: int,
    row: list[Any],
    header_map: dict[str, int],
) -> tuple[dict[str, Any] | None, str]:
    keyword = normalize_keyword(
        _cell_value(row, header_map, "keyword")
    )
    original_product_name = normalize_keyword(
        _cell_value(row, header_map, "original_product_name")
    )
    processed_product_name = normalize_keyword(
        _cell_value(row, header_map, "processed_product_name")
    )
    tag_cell = _cell_value(row, header_map, "tag_cell")
    tags = parse_tag_cell(tag_cell)

    missing_fields = []

    if not keyword:
        missing_fields.append("메인키워드")
    if not original_product_name:
        missing_fields.append("원본상품명")
    if not processed_product_name:
        missing_fields.append("가공상품명")
    if not tags:
        missing_fields.append("검색태그")

    if missing_fields:
        return None, "누락: " + ", ".join(missing_fields)

    seller_product_code = normalize_keyword(
        _cell_value(row, header_map, "seller_product_code")
    )
    sale_code = normalize_keyword(
        _cell_value(row, header_map, "sale_code")
    )
    category = normalize_keyword(
        _cell_value(row, header_map, "category")
    )

    fingerprint_payload = {
        "source_file_name": source_file.name,
        "sheet_name": sheet_name,
        "row_number": row_number,
        "seller_product_code": seller_product_code,
        "keyword": keyword,
        "processed_product_name": processed_product_name,
        "tags": tags,
    }

    return {
        "source_file_name": source_file.name,
        "source_file_path": str(source_file),
        "source_sheet": sheet_name,
        "source_row": row_number,
        "seller_product_code": seller_product_code,
        "sale_code": sale_code,
        "keyword": keyword,
        "original_product_name": original_product_name,
        "processed_product_name": processed_product_name,
        "category": category,
        "tags": tags,
        "fingerprint": stable_fingerprint(fingerprint_payload),
    }, ""



def _iter_openpyxl_rows(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise OptimizationWorkbookImportError(
            "openpyxl이 설치되어 있지 않습니다. "
            "'python -m pip install openpyxl'을 실행해 주세요."
        ) from error

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise OptimizationWorkbookImportError(
            f"엑셀파일을 열지 못했습니다: {path.name} / {error}"
        ) from error

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            headers_tuple = next(rows, None)

            if headers_tuple is None:
                yield worksheet.title, [], []
                continue

            headers = list(headers_tuple)
            data_rows = [
                list(row)
                for row in rows
            ]

            yield worksheet.title, headers, data_rows
    finally:
        workbook.close()



def _iter_xls_rows(path: Path):
    try:
        import pandas as pd
    except ImportError as error:
        raise OptimizationWorkbookImportError(
            ".xls 파일을 읽으려면 pandas와 xlrd가 필요합니다. "
            "'python -m pip install pandas xlrd'를 실행해 주세요."
        ) from error

    try:
        sheet_map = pd.read_excel(
            path,
            sheet_name=None,
            header=0,
            dtype=object,
        )
    except Exception as error:
        raise OptimizationWorkbookImportError(
            f".xls 파일을 열지 못했습니다: {path.name} / {error}"
        ) from error

    for sheet_name, dataframe in sheet_map.items():
        headers = list(dataframe.columns)
        data_rows = dataframe.where(
            dataframe.notna(),
            None,
        ).values.tolist()
        yield str(sheet_name), headers, data_rows



def analyze_optimization_workbooks(
    paths: list[Path],
) -> dict[str, Any]:
    records: list[dict[str, Any]] = []
    invalid_rows: list[dict[str, Any]] = []
    file_summaries: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for path in paths:
        file_valid_count = 0
        file_invalid_count = 0
        sheet_count = 0

        try:
            if path.suffix.lower() == ".xls":
                sheet_iterator = _iter_xls_rows(path)
            else:
                sheet_iterator = _iter_openpyxl_rows(path)

            for sheet_name, headers, data_rows in sheet_iterator:
                sheet_count += 1
                header_map = _build_header_map(headers)

                if not REQUIRED_FIELDS.issubset(header_map):
                    invalid_rows.append(
                        {
                            "source_file": path.name,
                            "source_sheet": sheet_name,
                            "source_row": 1,
                            "reason": (
                                "필수 열 없음: "
                                + ", ".join(
                                    sorted(REQUIRED_FIELDS - set(header_map))
                                )
                            ),
                        }
                    )
                    file_invalid_count += 1
                    continue

                for offset, row in enumerate(data_rows, start=2):
                    if not any(
                        value not in (None, "")
                        for value in row
                    ):
                        continue

                    record, reason = _build_record(
                        source_file=path,
                        sheet_name=sheet_name,
                        row_number=offset,
                        row=row,
                        header_map=header_map,
                    )

                    if record is None:
                        invalid_rows.append(
                            {
                                "source_file": path.name,
                                "source_sheet": sheet_name,
                                "source_row": offset,
                                "reason": reason,
                            }
                        )
                        file_invalid_count += 1
                        continue

                    records.append(record)
                    file_valid_count += 1

        except OptimizationWorkbookImportError as error:
            errors.append(
                {
                    "source_file": path.name,
                    "message": str(error),
                }
            )

        file_summaries.append(
            {
                "source_file": path.name,
                "sheet_count": sheet_count,
                "valid_record_count": file_valid_count,
                "invalid_row_count": file_invalid_count,
            }
        )

    unique_keywords = sorted(
        {
            record["keyword"]
            for record in records
        }
    )
    unique_tags = sorted(
        {
            tag["text"]
            for record in records
            for tag in record.get("tags", [])
        }
    )

    return {
        "files": [str(path) for path in paths],
        "file_summaries": file_summaries,
        "records": records,
        "valid_record_count": len(records),
        "invalid_rows": invalid_rows,
        "invalid_row_count": len(invalid_rows),
        "errors": errors,
        "error_count": len(errors),
        "unique_keywords": unique_keywords,
        "unique_keyword_count": len(unique_keywords),
        "unique_tags": unique_tags,
        "unique_tag_count": len(unique_tags),
    }

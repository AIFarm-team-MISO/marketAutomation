# utils/web_automation/smartstore_review_checker/review_result_writer.py

"""
스마트스토어 리뷰 필터링 결과 엑셀 저장 모듈

역할:
- 리뷰 매칭 결과를 하나의 엑셀 파일로 저장한다.
- 비교상품 파일과 같은 폴더에 결과 파일을 생성한다.
- 결과 시트를 요약 / 전체 / 리뷰있음 / 리뷰없음 / 리뷰3개이상 / 리뷰1~2개로 나눈다.
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.global_logger import logger


def sanitize_file_name(file_name):
    """
    윈도우 파일명에서 사용할 수 없는 문자를 제거한다.
    """

    invalid_chars = [
        "\\",
        "/",
        ":",
        "*",
        "?",
        '"',
        "<",
        ">",
        "|",
    ]

    cleaned_file_name = file_name

    for invalid_char in invalid_chars:
        cleaned_file_name = cleaned_file_name.replace(
            invalid_char,
            "_",
        )

    return cleaned_file_name


def make_review_result_output_file_path(
    output_folder_path,
    market_display_name,
):
    """
    리뷰 필터링 결과 엑셀 저장 경로를 만든다.
    """

    os.makedirs(
        output_folder_path,
        exist_ok=True,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    file_name = sanitize_file_name(
        f"리뷰필터링완료_output_{market_display_name}_{timestamp}.xlsx"
    )

    output_file_path = os.path.join(
        output_folder_path,
        file_name,
    )

    return output_file_path


def make_result_summary_dataframe(
    market_display_name,
    master_file_path,
    target_file_path,
    output_file_path,
    matched_dataframe,
):
    """
    결과 요약 시트를 만든다.
    """

    total_count = len(matched_dataframe)

    review_exists_count = int(
        (matched_dataframe["총리뷰수"] >= 1).sum()
    )

    no_review_count = int(
        (matched_dataframe["총리뷰수"] == 0).sum()
    )

    review_over_3_count = int(
        (matched_dataframe["총리뷰수"] >= 3).sum()
    )

    review_1_to_2_count = int(
        (
            (matched_dataframe["총리뷰수"] >= 1)
            & (matched_dataframe["총리뷰수"] <= 2)
        ).sum()
    )

    summary_rows = [
        {
            "항목": "마켓명",
            "값": market_display_name,
        },
        {
            "항목": "생성일시",
            "값": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        {
            "항목": "리뷰통합본파일",
            "값": master_file_path,
        },
        {
            "항목": "비교대상파일",
            "값": target_file_path,
        },
        {
            "항목": "결과파일",
            "값": output_file_path,
        },
        {
            "항목": "전체비교상품수",
            "값": total_count,
        },
        {
            "항목": "리뷰있음",
            "값": review_exists_count,
        },
        {
            "항목": "리뷰없음",
            "값": no_review_count,
        },
        {
            "항목": "리뷰3개이상",
            "값": review_over_3_count,
        },
        {
            "항목": "리뷰1~2개",
            "값": review_1_to_2_count,
        },
        {
            "항목": "분류기준",
            "값": "상품번호 기준 매칭 후 총리뷰수로 분류",
        },
    ]

    return pd.DataFrame(summary_rows)


def format_result_excel_file(excel_file_path):
    """
    생성된 결과 엑셀 파일의 기본 서식을 정리한다.
    """

    workbook = load_workbook(excel_file_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True,
    )

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = thin_border

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=False,
                )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                value_length = len(str(cell.value))

                if value_length > max_length:
                    max_length = value_length

            adjusted_width = min(
                max(max_length + 2, 10),
                45,
            )

            worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_file_path)


def write_review_filter_result_excel(
    output_folder_path,
    market_display_name,
    master_file_path,
    target_file_path,
    matched_dataframe,
    split_dataframes,
):
    """
    리뷰 필터링 결과 엑셀 파일을 저장한다.
    """

    output_file_path = make_review_result_output_file_path(
        output_folder_path=output_folder_path,
        market_display_name=market_display_name,
    )

    summary_dataframe = make_result_summary_dataframe(
        market_display_name=market_display_name,
        master_file_path=master_file_path,
        target_file_path=target_file_path,
        output_file_path=output_file_path,
        matched_dataframe=matched_dataframe,
    )

    with pd.ExcelWriter(
        output_file_path,
        engine="openpyxl",
    ) as writer:
        summary_dataframe.to_excel(
            writer,
            sheet_name="00_요약",
            index=False,
        )

        split_dataframes["all"].to_excel(
            writer,
            sheet_name="01_전체비교결과",
            index=False,
        )

        split_dataframes["review_exists"].to_excel(
            writer,
            sheet_name="02_리뷰있음",
            index=False,
        )

        split_dataframes["no_review"].to_excel(
            writer,
            sheet_name="03_리뷰없음",
            index=False,
        )

        split_dataframes["review_over_3"].to_excel(
            writer,
            sheet_name="04_리뷰3개이상",
            index=False,
        )

        split_dataframes["review_1_to_2"].to_excel(
            writer,
            sheet_name="05_리뷰1~2개",
            index=False,
        )

    format_result_excel_file(output_file_path)

    logger.log(
        (
            "[리뷰 필터링 결과 엑셀 저장 완료]\n"
            f"- 파일: {output_file_path}\n"
            f"- 전체비교상품: {len(split_dataframes['all'])}개\n"
            f"- 리뷰있음: {len(split_dataframes['review_exists'])}개\n"
            f"- 리뷰없음: {len(split_dataframes['no_review'])}개\n"
            f"- 리뷰3개이상: {len(split_dataframes['review_over_3'])}개\n"
            f"- 리뷰1~2개: {len(split_dataframes['review_1_to_2'])}개"
        ),
        level="INFO",
        also_to_report=True,
        emoji_key="스마트스토어",
        separator="1line",
    )

    return output_file_path
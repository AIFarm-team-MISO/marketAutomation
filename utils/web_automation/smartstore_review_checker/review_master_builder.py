# utils/web_automation/smartstore_review_checker/review_master_builder.py

"""
스마트스토어 리뷰 통합본 생성 모듈

역할:
- review_data/{market_name} 폴더 안의 리뷰 엑셀 원본들을 읽는다.
- 전시상태가 정상인 리뷰만 기본 집계에 사용한다.
- 상품번호 기준으로 리뷰 집계를 만든다.
- 정상 리뷰 전체 행을 통합정상리뷰 시트로 저장한다.
- 블라인드 리뷰는 집계에서 제외하고 확인필요 시트로 따로 저장한다.

생성 파일:
- review_data/{market_name}/master/{market_name}_review_master.xlsx
"""

import os
import warnings
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.global_logger import logger

from utils.web_automation.smartstore_review_checker.review_download_reader import (
    read_review_download_folder,
)


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style*",
)


CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

REVIEW_DATA_DIR = os.path.join(
    CURRENT_DIR,
    "review_data",
)


def get_market_review_source_folder_path(market_name):
    """
    마켓별 리뷰 원본 엑셀 폴더 경로를 반환한다.

    중요:
    - 원본 리뷰 폴더는 자동 생성하지 않는다.
    - 마켓 폴더는 사용자가 직접 만든다.

    이유:
    - 마켓키 오타가 있을 때 잘못된 폴더가 자동 생성되는 것을 막기 위함이다.
    - 예: today_dameum을 today_damum으로 잘못 입력했을 때
      잘못된 폴더가 생기면 실수를 알아차리기 어렵다.

    예:
    market_name = "onestop_living"
    → review_data/onestop_living

    market_name = "today_dameum"
    → review_data/today_dameum
    """

    source_folder_path = os.path.join(
        REVIEW_DATA_DIR,
        market_name,
    )

    return source_folder_path

def validate_market_review_source_folder(
    market_name,
    market_display_name,
    source_folder_path,
):
    """
    마켓별 리뷰 원본 폴더가 실제로 존재하는지 확인한다.

    원본 폴더는 자동 생성하지 않는다.
    없으면 작업을 중단하고, 사용자가 직접 만들도록 안내한다.
    """

    if os.path.exists(source_folder_path):
        return

    logger.log(
        (
            "[리뷰 원본 폴더 없음]\n"
            f"- 마켓: {market_display_name}\n"
            f"- 마켓키: {market_name}\n"
            f"- 필요한 폴더: {source_folder_path}\n"
            "- 위 폴더를 직접 만든 뒤, 스마트스토어 리뷰 다운로드 엑셀 파일을 넣고 다시 실행하세요."
        ),
        level="ERROR",
        also_to_report=True,
        emoji_key="스마트스토어",
        separator="1line",
    )

    raise FileNotFoundError(
        f"리뷰 원본 폴더가 없습니다: {source_folder_path}"
    )

def get_review_source_file_names(source_folder_path):
    """
    리뷰 원본 폴더 안에서 읽을 수 있는 리뷰 엑셀 파일명 목록을 반환한다.

    제외 대상:
    - 임시 엑셀 파일: ~$ 로 시작하는 파일
    - 결과 파일: 리뷰필터링완료_output 포함 파일
    - master 파일: _review_master 포함 파일
    """

    allowed_extensions = (
        ".xlsx",
        ".xls",
    )

    exclude_keywords = [
        "리뷰필터링완료_output",
        "review_filter_output",
        "_review_master",
    ]

    file_names = []

    for file_name in os.listdir(source_folder_path):
        if file_name.startswith("~$"):
            continue

        if any(keyword in file_name for keyword in exclude_keywords):
            continue

        file_path = os.path.join(
            source_folder_path,
            file_name,
        )

        if not os.path.isfile(file_path):
            continue

        _, file_extension = os.path.splitext(file_name)
        file_extension = file_extension.lower()

        if file_extension in allowed_extensions:
            file_names.append(file_name)

    file_names.sort()

    return file_names


def get_market_review_master_folder_path(market_name):
    """
    마켓별 리뷰 master 저장 폴더 경로를 반환한다.

    중요:
    - 원본 리뷰 폴더는 여기서 만들지 않는다.
    - 원본 리뷰 폴더가 이미 존재할 때만 master 폴더를 자동 생성한다.

    예:
    review_data/today_dameum          ← 사용자가 직접 생성
    review_data/today_dameum/master   ← 코드가 자동 생성
    """

    source_folder_path = get_market_review_source_folder_path(
        market_name
    )

    if not os.path.exists(source_folder_path):
        raise FileNotFoundError(
            f"리뷰 원본 폴더가 없어 master 폴더를 만들 수 없습니다: {source_folder_path}"
        )

    master_folder_path = os.path.join(
        source_folder_path,
        "master",
    )

    os.makedirs(
        master_folder_path,
        exist_ok=True,
    )

    return master_folder_path


def get_market_review_master_file_path(market_name):
    """
    마켓별 리뷰 통합본 엑셀 파일 경로를 반환한다.
    """

    master_folder_path = get_market_review_master_folder_path(
        market_name
    )

    return os.path.join(
        master_folder_path,
        f"{market_name}_review_master.xlsx",
    )


def clean_text_value(value):
    """
    값을 문자열로 정리한다.
    """

    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() in ["nan", "none", "nat"]:
        return ""

    return text


def add_review_datetime_column(review_dataframe):
    """
    리뷰등록일 문자열을 datetime 컬럼으로 변환한다.
    """

    review_dataframe = review_dataframe.copy()

    review_dataframe["리뷰등록일_datetime"] = pd.to_datetime(
        review_dataframe["리뷰등록일"],
        format="%Y.%m.%d. %H:%M:%S",
        errors="coerce",
    )

    return review_dataframe


def filter_reviews_by_display_status(review_dataframe):
    """
    전시상태 기준으로 정상 리뷰와 확인필요 리뷰를 나눈다.

    정상 리뷰:
    - 상품별 리뷰 집계에 사용

    확인필요 리뷰:
    - 블라인드 등 정상 상태가 아닌 리뷰
    - 집계에서는 제외
    - 별도 시트로 저장
    """

    normal_review_dataframe = review_dataframe[
        review_dataframe["전시상태"] == "정상"
    ].copy()

    need_check_review_dataframe = review_dataframe[
        review_dataframe["전시상태"] != "정상"
    ].copy()

    return normal_review_dataframe, need_check_review_dataframe


def prepare_review_dataframe_for_aggregation(review_dataframe):
    """
    상품번호별 집계를 위한 보조 컬럼을 만든다.
    """

    review_dataframe = review_dataframe.copy()

    review_dataframe = add_review_datetime_column(review_dataframe)

    review_dataframe["구매자평점_number"] = pd.to_numeric(
        review_dataframe["구매자평점"],
        errors="coerce",
    )

    review_dataframe["일반리뷰여부"] = (
        review_dataframe["리뷰구분"] == "일반"
    ).astype(int)

    review_dataframe["한달사용리뷰여부"] = (
        review_dataframe["리뷰구분"] == "한달사용"
    ).astype(int)

    review_dataframe["저평점리뷰여부"] = (
        review_dataframe["구매자평점_number"] <= 2
    ).astype(int)

    if "포토/영상" in review_dataframe.columns:
        review_dataframe["포토리뷰여부"] = (
            review_dataframe["포토/영상"]
            .apply(clean_text_value)
            .apply(lambda value: 1 if value else 0)
        )
    else:
        review_dataframe["포토리뷰여부"] = 0

    if "베스트리뷰" in review_dataframe.columns:
        review_dataframe["베스트리뷰여부"] = (
            review_dataframe["베스트리뷰"]
            .apply(clean_text_value)
            .apply(lambda value: 1 if value == "Y" else 0)
        )
    else:
        review_dataframe["베스트리뷰여부"] = 0

    return review_dataframe


def make_product_review_summary(normal_review_dataframe):
    """
    정상 리뷰 DataFrame을 상품번호 기준으로 집계한다.

    여기서는 리뷰글번호 기준 중복 제거를 하지 않는다.
    정상 리뷰 행 전체를 상품번호별로 집계한다.
    """

    review_dataframe = prepare_review_dataframe_for_aggregation(
        normal_review_dataframe
    )

    review_dataframe = review_dataframe.sort_values(
        by=[
            "상품번호",
            "리뷰등록일_datetime",
        ],
        ascending=[
            True,
            True,
        ],
    )

    product_summary_dataframe = (
        review_dataframe
        .groupby("상품번호", as_index=False)
        .agg(
            대표상품명=("상품명", "last"),
            총리뷰수=("상품번호", "size"),
            일반리뷰수=("일반리뷰여부", "sum"),
            한달사용리뷰수=("한달사용리뷰여부", "sum"),
            평균평점=("구매자평점_number", "mean"),
            최저평점=("구매자평점_number", "min"),
            최고평점=("구매자평점_number", "max"),
            저평점리뷰수=("저평점리뷰여부", "sum"),
            포토리뷰수=("포토리뷰여부", "sum"),
            베스트리뷰수=("베스트리뷰여부", "sum"),
            최초리뷰일=("리뷰등록일_datetime", "min"),
            최근리뷰일=("리뷰등록일_datetime", "max"),
        )
    )

    product_summary_dataframe["평균평점"] = (
        product_summary_dataframe["평균평점"]
        .round(2)
    )

    product_summary_dataframe["최초리뷰일"] = (
        product_summary_dataframe["최초리뷰일"]
        .dt.strftime("%Y-%m-%d %H:%M:%S")
        .fillna("")
    )

    product_summary_dataframe["최근리뷰일"] = (
        product_summary_dataframe["최근리뷰일"]
        .dt.strftime("%Y-%m-%d %H:%M:%S")
        .fillna("")
    )

    product_summary_dataframe = product_summary_dataframe.sort_values(
        by=[
            "총리뷰수",
            "최근리뷰일",
        ],
        ascending=[
            False,
            False,
        ],
    )

    return product_summary_dataframe


def make_master_summary_dataframe(
    market_name,
    market_display_name,
    source_folder_path,
    master_file_path,
    original_review_dataframe,
    normal_review_dataframe,
    need_check_review_dataframe,
    product_summary_dataframe,
):
    """
    통합본 요약 시트를 만든다.
    """

    summary_rows = [
        {
            "항목": "마켓키",
            "값": market_name,
        },
        {
            "항목": "마켓명",
            "값": market_display_name,
        },
        {
            "항목": "생성일시",
            "값": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        {
            "항목": "원본폴더",
            "값": source_folder_path,
        },
        {
            "항목": "통합본파일",
            "값": master_file_path,
        },
        {
            "항목": "원본리뷰행수",
            "값": len(original_review_dataframe),
        },
        {
            "항목": "정상리뷰행수",
            "값": len(normal_review_dataframe),
        },
        {
            "항목": "확인필요리뷰행수",
            "값": len(need_check_review_dataframe),
        },
        {
            "항목": "상품별집계상품수",
            "값": len(product_summary_dataframe),
        },
        {
            "항목": "집계기준",
            "값": "전시상태 정상 리뷰 전체를 상품번호 기준으로 집계",
        },
        {
            "항목": "제외기준",
            "값": "전시상태가 정상이 아닌 리뷰는 확인필요 시트로 분리",
        },
    ]

    return pd.DataFrame(summary_rows)


def format_review_excel_file(excel_file_path):
    """
    생성된 리뷰 통합본 엑셀 파일의 기본 서식을 정리한다.
    """

    workbook = load_workbook(excel_file_path)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True,
    )

    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"

        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = thin_border

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=False,
                )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue

                value_length = len(str(cell.value))

                if value_length > max_length:
                    max_length = value_length

            adjusted_width = min(
                max(max_length + 2, 10),
                45,
            )

            worksheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_file_path)


def write_review_master_excel(
    master_file_path,
    summary_dataframe,
    product_summary_dataframe,
    normal_review_dataframe,
    need_check_review_dataframe,
):
    """
    리뷰 통합본 엑셀 파일을 저장한다.
    """

    with pd.ExcelWriter(
        master_file_path,
        engine="openpyxl",
    ) as writer:
        summary_dataframe.to_excel(
            writer,
            sheet_name="00_요약",
            index=False,
        )

        product_summary_dataframe.to_excel(
            writer,
            sheet_name="01_상품별리뷰집계",
            index=False,
        )

        normal_review_dataframe.to_excel(
            writer,
            sheet_name="02_통합정상리뷰",
            index=False,
        )

        if not need_check_review_dataframe.empty:
            need_check_review_dataframe.to_excel(
                writer,
                sheet_name="03_확인필요리뷰",
                index=False,
            )

    format_review_excel_file(master_file_path)


def build_review_master(
    market_name,
    market_display_name=None,
):
    """
    마켓별 리뷰 원본 엑셀을 통합하여 리뷰 master 엑셀을 생성한다.

    처리 흐름:
    1. market_name 기준으로 리뷰 원본 폴더 경로를 가져온다.
       - 원본 폴더는 자동 생성하지 않는다.
       - 사용자가 직접 만들어야 한다.

    2. 원본 폴더가 실제로 존재하는지 확인한다.
       - 없으면 안내 로그를 출력하고 작업을 중단한다.

    3. 원본 폴더 안에 리뷰 엑셀 파일이 있는지 확인한다.
       - 없으면 안내 로그를 출력하고 작업을 중단한다.

    4. 원본 폴더가 정상일 때만 master 폴더와 master 파일 경로를 만든다.
       - master 폴더는 코드가 자동 생성한다.

    5. 리뷰 엑셀들을 통합하고,
       전시상태 정상 리뷰만 상품번호 기준으로 집계한다.

    6. master 엑셀을 저장한다.

    market_name 예:
    - onestop_living
    - today_dameum

    market_display_name 예:
    - 원스톱리빙
    - 오늘담음
    """

    # market_display_name이 따로 전달되지 않은 경우에만
    # market_name을 화면 표시용 이름으로 사용한다.
    if market_display_name is None:
        market_display_name = market_name

    # 1. 마켓별 리뷰 원본 폴더 경로를 가져온다.
    #    여기서는 폴더를 자동 생성하지 않는다.
    #    예: review_data/today_dameum
    source_folder_path = get_market_review_source_folder_path(
        market_name
    )

    # 2. 원본 폴더가 실제로 존재하는지 확인한다.
    #    없으면 사용자가 직접 만들도록 안내하고 작업을 중단한다.
    validate_market_review_source_folder(
        market_name=market_name,
        market_display_name=market_display_name,
        source_folder_path=source_folder_path,
    )

    # 3. 원본 폴더 안에 리뷰 엑셀 파일이 있는지 확인한다.
    #    폴더는 있지만 리뷰 파일을 아직 넣지 않은 경우를 잡기 위함이다.
    review_source_file_names = get_review_source_file_names(
        source_folder_path
    )

    if not review_source_file_names:
        logger.log(
            (
                "[리뷰 원본 파일 없음]\n"
                f"- 마켓: {market_display_name}\n"
                f"- 마켓키: {market_name}\n"
                f"- 리뷰 원본 폴더: {source_folder_path}\n"
                "- 위 폴더에 스마트스토어 리뷰 다운로드 엑셀 파일을 넣은 뒤 다시 실행하세요."
            ),
            level="ERROR",
            also_to_report=True,
            emoji_key="스마트스토어",
            separator="1line",
        )

        raise FileNotFoundError(
            f"리뷰 원본 엑셀 파일이 없습니다: {source_folder_path}"
        )

    # 4. 원본 폴더가 존재할 때만 master 폴더와 master 파일 경로를 만든다.
    #    master 폴더는 코드가 자동 생성한다.
    master_file_path = get_market_review_master_file_path(
        market_name
    )

    logger.log(
        (
            "[리뷰 통합본 생성 시작]\n"
            f"- 마켓: {market_display_name}\n"
            f"- 마켓키: {market_name}\n"
            f"- 원본폴더: {source_folder_path}\n"
            f"- 원본파일수: {len(review_source_file_names)}개"
        ),
        level="INFO",
        also_to_report=True,
        emoji_key="스마트스토어",
        separator="1line",
    )

    # 5. 마켓별 리뷰 원본 폴더 안의 엑셀 파일들을 통합한다.
    original_review_dataframe = read_review_download_folder(
        source_folder_path
    )

    # 6. 전시상태 기준으로 정상 리뷰와 확인필요 리뷰를 분리한다.
    normal_review_dataframe, need_check_review_dataframe = (
        filter_reviews_by_display_status(original_review_dataframe)
    )

    # 7. 정상 리뷰만 상품번호 기준으로 집계한다.
    product_summary_dataframe = make_product_review_summary(
        normal_review_dataframe
    )

    # 8. master 요약 시트를 만든다.
    summary_dataframe = make_master_summary_dataframe(
        market_name=market_name,
        market_display_name=market_display_name,
        source_folder_path=source_folder_path,
        master_file_path=master_file_path,
        original_review_dataframe=original_review_dataframe,
        normal_review_dataframe=normal_review_dataframe,
        need_check_review_dataframe=need_check_review_dataframe,
        product_summary_dataframe=product_summary_dataframe,
    )

    # 9. master 엑셀 파일을 저장한다.
    write_review_master_excel(
        master_file_path=master_file_path,
        summary_dataframe=summary_dataframe,
        product_summary_dataframe=product_summary_dataframe,
        normal_review_dataframe=normal_review_dataframe,
        need_check_review_dataframe=need_check_review_dataframe,
    )

    logger.log(
        (
            "[리뷰 통합본 생성 완료]\n"
            f"- 마켓: {market_display_name}\n"
            f"- 마켓키: {market_name}\n"
            f"- 통합본: {master_file_path}\n"
            f"- 원본리뷰: {len(original_review_dataframe)}개\n"
            f"- 정상리뷰: {len(normal_review_dataframe)}개\n"
            f"- 확인필요: {len(need_check_review_dataframe)}개\n"
            f"- 집계상품: {len(product_summary_dataframe)}개"
        ),
        level="INFO",
        also_to_report=True,
        emoji_key="스마트스토어",
        separator="1line",
    )

    return {
        "market_name": market_name,
        "market_display_name": market_display_name,
        "source_folder_path": source_folder_path,
        "master_file_path": master_file_path,
        "summary_dataframe": summary_dataframe,
        "product_summary_dataframe": product_summary_dataframe,
        "normal_review_dataframe": normal_review_dataframe,
        "need_check_review_dataframe": need_check_review_dataframe,
    }
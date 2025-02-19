from utils.global_logger import logger

from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd

'''
    이곳의 함수들은 모두 
    .xls 파일을 읽어 데이터를 처리한 후 .xlsx 형식으로 저장하는 함수로서 사용됨

    엑셀 편집과 관련된 유틸들 

'''

def update_excel_column(sheets, column_name, content, sheet_name=None):

    """
        해당경로의 엑셀파일을 읽어 지정된 칼럼의 내용을 변경후 기존 엑셀파일과 동일하게 만들어 저장
    
    Parameters:
        sheets (dict): 시트 이름을 키로, DataFrame을 값으로 갖는 딕셔너리.
        sheet_name (str): 수정할 시트 이름.
        column_name (str): 추가하거나 업데이트할 열 이름.
        content (str): 열에 입력할 내용.
        sheet_name (str): 엑셀파일의 시트중 변경할 시트명, 없으면 첫번째 시트사용

        Returns:
        DataFrame: 수정된 DataFrame
        
    """
    try:


        # 시트를 지정하지 않으면 첫 번째 시트 가져오기
        if sheet_name is None:
            sheet_name = list(sheets.keys())[0]

        df = sheets[sheet_name]

        logger.log(f" [{column_name}] 열을 {content}로 변경합니다. ", level="INFO")

        # 기존의 열을 문자열로 변환 (NaN 값을 빈 문자열로 초기화 후 문자열로 변환)
        df[column_name] = df[column_name].fillna("").astype(str)

        # 2행부터 데이터 업데이트
        df.loc[1:, column_name] = content

        return df
    
    except Exception as e:
        logger.log(f"오류 발생 - 시트: {sheet_name}, 이유: {e}", level="ERROR")


def insert_excel_column(first_sheet_data, existing_column_name, new_column_name, position="before", offset=1):
    """
    지정된 열 이름의 앞이나 뒤에 새 열을 삽입하고 3행부터 빈 문자열로 초기화한 DataFrame을 반환.

    Parameters:
        
        first_sheet_data : 엑셀파일의 시트 중 첫번째 시트
        existing_column_name (str): 새 열을 삽입할 기준 열 이름.
        new_column_name (str): 삽입할 새 열 이름.
        position (str): 'before' 또는 'after'를 지정하여 열 삽입 위치 결정.
        offset (int): 기준 열로부터의 거리.

    Returns:
        DataFrame: 수정된 DataFrame
    """
    try:

        df = first_sheet_data

        if existing_column_name not in df.columns:
            raise ValueError(f"기준 열 '{existing_column_name}'이(가) 존재하지 않습니다.")
    

        # 기준 열 위치 확인
        base_idx = df.columns.get_loc(existing_column_name)

        # 삽입 위치 계산
        if position == "before":
            insert_idx = max(0, base_idx - offset + 1)
        elif position == "after":
            insert_idx = min(len(df.columns), base_idx + offset)
        else:
            raise ValueError("position은 'before' 또는 'after' 중 하나여야 합니다.")

        # 새로운 열 삽입 (3행부터 빈 문자열로 초기화)
        new_column_data = ["" if i >= 2 else None for i in range(len(df))]
        df.insert(loc=insert_idx, column=new_column_name, value=new_column_data)

        logger.log(f"'{new_column_name}' 열을 '{existing_column_name}'의 {position} 위치에 삽입했습니다.")


        return df

    except Exception as e:
        logger.log(f"열 삽입 중 오류 발생: {e}", level="ERROR")
        raise

def insert_excel_multiindex_column(first_sheet_data, existing_column_name, new_column_tuple, position="before", offset=1):
    """
    멀티 인덱스(더블 컬럼)에서 지정된 열의 앞이나 뒤에 새로운 컬럼을 삽입하는 함수.
    새 컬럼은 3행부터 빈 문자열("")로 초기화됨.

    Parameters:
        first_sheet_data : DataFrame (멀티 인덱스 컬럼 포함 가능)
        existing_column_name (str): 기준이 되는 첫 번째 레벨의 컬럼명.
        new_column_tuple (tuple): 삽입할 새 컬럼 (예: ("새 컬럼 그룹", "새 컬럼 이름"))
        position (str): 'before' 또는 'after'를 지정하여 열 삽입 위치 결정.
        offset (int): 기준 열로부터의 거리.

    Returns:
        DataFrame: 수정된 DataFrame
    """
    try:
        df = first_sheet_data.copy()

        if not isinstance(df.columns, pd.MultiIndex):
            raise ValueError("데이터프레임의 컬럼이 멀티 인덱스가 아닙니다. 멀티 인덱스가 필요합니다.")

        # 첫 번째 레벨(메인 컬럼)에서 기준 열의 위치 찾기
        first_level_columns = df.columns.get_level_values(0)
        
        if existing_column_name not in first_level_columns:
            raise ValueError(f"기준 열 '{existing_column_name}'을 찾을 수 없습니다.")

        base_idx = list(first_level_columns).index(existing_column_name)

        # 삽입 위치 계산
        if position == "before":
            insert_idx = max(0, base_idx - offset + 1)
        elif position == "after":
            insert_idx = min(len(df.columns), base_idx + offset)
        else:
            raise ValueError("position은 'before' 또는 'after' 중 하나여야 합니다.")

        # 새로운 열 삽입 (3행부터 빈 문자열로 초기화)
        new_column_data = ["" if i >= 2 else None for i in range(len(df))]
        
        # 새로운 컬럼 추가 (멀티 인덱스를 고려하여 튜플 형태로 추가)
        new_columns = list(df.columns)
        new_columns.insert(insert_idx, new_column_tuple)

        df = df.reindex(columns=new_columns)  # 컬럼 재정렬
        df[new_column_tuple] = new_column_data  # 새 컬럼 데이터 삽입

        logger.log(f"'{new_column_tuple}' 열을 '{existing_column_name}'의 {position} 위치에 삽입했습니다.", level="INFO")

        return df

    except Exception as e:
        raise ValueError(f"멀티 인덱스에서 열 삽입 중 문제가 발생했습니다: {e}")
    

def insert_excel_columns_with_values(df, existing_column_name, filter_image_column_name, filtered_result_column_name):
    """
    멀티 인덱스(더블 컬럼) 및 일반 컬럼 모두에서 
    '필터링결과'와 '썸네일이미지' 컬럼을 삽입하고,
    '썸네일이미지' 컬럼에 extract_main_image_urls() 함수에서 추출한 main URL을 입력하는 함수.

    Parameters:
        df (pd.DataFrame): 데이터프레임 (멀티 인덱스 또는 일반 인덱스)
        existing_column_name (str): 기준이 되는 컬럼명 (예: "이미지명")
        filter_image_column_name (str): 썸네일만 추출한 칼럼명
        filtered_result_column_name (str): 필터링후 결과값이 입력되는 컬럼명

    Returns:
        pd.DataFrame: 수정된 데이터프레임
    """
    try:
        # ✅ 멀티 인덱스 여부 확인
        is_multiindex = isinstance(df.columns, pd.MultiIndex)

        # ✅ 기준 열 존재 여부 확인
        if is_multiindex:
            first_level_columns = df.columns.get_level_values(0)
            if existing_column_name not in first_level_columns:
                raise KeyError(f"❌ 멀티 인덱스에서 '{existing_column_name}' 컬럼을 찾을 수 없습니다.")
            base_idx = list(first_level_columns).index(existing_column_name)
        else:
            if existing_column_name not in df.columns:
                raise KeyError(f"❌ 일반 컬럼에서 '{existing_column_name}'을 찾을 수 없습니다.")
            base_idx = df.columns.get_loc(existing_column_name)

        # ✅ main 이미지 URL 추출
        main_image_list = extract_main_image_urls(df, column_name=existing_column_name)

        # ✅ 삽입할 컬럼과 기본값 설정
        new_columns_with_values = {
            (filtered_result_column_name, "result") if is_multiindex else filtered_result_column_name: "",  # 빈 값
            (filter_image_column_name, "sum_image") if is_multiindex else filter_image_column_name: main_image_list  # main 이미지 URL
        }

        # ✅ 기존 df 복사
        updated_df = df.copy()

        # ✅ 컬럼 삽입
        for col, value in reversed(new_columns_with_values.items()):
            col_name = col if is_multiindex else str(col)
            updated_df.insert(base_idx, col_name, value)

        logger.log(f"✅ '{existing_column_name}' 앞에 '필터링결과'와 '썸네일이미지' 컬럼 추가 완료.", level="INFO")

        return updated_df

    except Exception as e:
        raise ValueError(f"❌ 열 삽입 및 값 입력 중 문제가 발생했습니다: {e}")


    

def extract_main_image_urls(dataframe, column_name="이미지명"):
    """
    '이미지명' 열에서 'main' 이후의 이미지 URL을 추출하여 리스트로 반환하는 함수.
    
    ✅ 멀티 인덱스 & 단일 인덱스 모두 지원.
    
    Parameters:
        dataframe (pd.DataFrame): 데이터프레임 (멀티 인덱스 또는 일반 인덱스).
        column_name (str): 이미지 URL이 저장된 컬럼명 (기본값: "이미지명").

    Returns:
        list: 'main' 이후의 이미지 URL 리스트.
    """
    try:
        extracted_urls = []

        # ✅ 멀티 인덱스 여부 확인
        is_multiindex = isinstance(dataframe.columns, pd.MultiIndex)

        # ✅ 컬럼 찾기 (멀티 인덱스일 경우 첫 번째 레벨 기준)
        if is_multiindex:
            first_level_columns = dataframe.columns.get_level_values(0)
            if column_name not in first_level_columns:
                raise KeyError(f"❌ 멀티 인덱스에서 '{column_name}' 컬럼을 찾을 수 없습니다.")
            target_columns = [col for col in dataframe.columns if col[0] == column_name]
        else:
            if column_name not in dataframe.columns:
                raise KeyError(f"❌ 일반 컬럼에서 '{column_name}'을 찾을 수 없습니다.")
            target_columns = [column_name]

        # ✅ 데이터프레임의 각 행에서 'main' 이후의 URL 추출
        for _, row in dataframe[target_columns].dropna(how="all").iterrows():
            main_url = None

            for col in target_columns:
                value = row[col]
                if isinstance(value, str):  # 문자열이 아닌 경우 스킵
                    sections = value.split("\n")  # 줄바꿈 기준으로 나누기
                    for section in sections:
                        if section.startswith("main^|^"):
                            main_url = section.split("^|^")[1]  # 'main^|^' 이후의 URL 추출
                            break
                if main_url:
                    extracted_urls.append(main_url)
                    break

        logger.log(f"✅ '{column_name}'에서 main 이미지 URL {len(extracted_urls)}개 추출 완료.", level="INFO")
        return extracted_urls

    except Exception as e:
        raise ValueError(f"❌ 이미지명 칼럼에서 main URL을 추출하는 중 문제가 발생했습니다: {e}")







def clean_data(sheet):
    """
    데이터프레임에서 빈 행을 제거하고 인덱스를 초기화합니다.
    """
    sheet = sheet.dropna(how="all")  # 모든 값이 NaN인 행 제거
    sheet.reset_index(drop=True, inplace=True)  # 인덱스 초기화
    return sheet

def sort_sheet(sheet, highlight_column, highlight_value):
    """
    특정 열의 값에 해당하는 행을 강조하고, 해당 행을 맨 위로 정렬합니다.

    Parameters:
        sheet (pd.DataFrame): 편집할 DataFrame.
        highlight_column (str): 강조할 열 이름.
        highlight_value (str): 강조 조건 값.

    Returns:
        pd.DataFrame: 강조 및 정렬된 DataFrame.
    """
    try:
        logger.log("정렬 작업 시작", level="INFO")

        sheet = clean_data(sheet)

        # 조건에 맞는 행과 그렇지 않은 행으로 분리
        highlight_rows = sheet[sheet[highlight_column] == highlight_value]
        other_rows = sheet[sheet[highlight_column] != highlight_value]

        # 강조 및 비강조 행 수 확인
        logger.log(f"강조 행 갯수 ({highlight_value}): {len(highlight_rows)}", level="INFO")
        logger.log(f"비강조 행 갯수: {len(other_rows)}", level="INFO")
        logger.log(f"전체 행 갯수: {len(sheet)}", level="INFO")

        # 강조 행을 맨 위로 정렬
        if highlight_rows.empty:
            logger.log(f"강조 행 ({highlight_value})가 없습니다. 기존 순서를 유지합니다.", level="WARNING")
            return sheet  # 정렬하지 않고 원래 DataFrame 반환

        # 강조 행을 맨 위로 정렬
        sorted_sheet = pd.concat([highlight_rows, other_rows], ignore_index=True)

         # 정렬 결과 확인
        logger.log(f"강조 행 ({highlight_value})이(가) 맨 위로 정렬되었습니다.", level="INFO")


        return sorted_sheet

    except Exception as e:
        logger.log(f"정렬 작업 중 오류 발생: {e}", level="ERROR")
        raise




from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def colum_highlight_sheet(sheets, filtered_sort_sheets, highlight_column, highlight_value, output_file_name):
    """
    모든 시트를 유지하면서 정렬된 시트를 반영하고, 강조 작업을 수행한 후 저장.

    Parameters:
        sheets (dict): 원래 엑셀 파일의 모든 시트 데이터 {시트 이름: DataFrame}.
        filtered_sort_sheets (pd.DataFrame): 정렬된 첫 번째 시트 데이터.
        highlight_column (str): 강조할 열 이름.
        highlight_value (str): 강조 조건 값.
        output_file_path (str): 저장할 출력 파일 경로.

    Returns:
        None
    """
    try:
        # 워크북 생성
        workbook = Workbook()
        highlight_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for idx, (sheet_name, df) in enumerate(sheets.items()):
            if idx == 0:  # 첫 번째 시트는 정렬된 데이터로 대체
                worksheet = workbook.active
                worksheet.title = sheet_name

                # 헤더 추가
                headers = list(filtered_sort_sheets.columns)
                worksheet.append(headers)

                # 한 행 비우기
                worksheet.append([None] * len(headers))

                # 데이터 추가 및 강조 작업
                for i, row in filtered_sort_sheets.iterrows():
                    worksheet.append(row.tolist())
                    if highlight_column in filtered_sort_sheets.columns and row[highlight_column] == highlight_value:
                        for cell in worksheet[i + 3]:  # 헤더(1행) + 빈 행(1행) + 데이터 행 보정(+3)
                            cell.fill = highlight_fill
            else:  # 나머지 시트는 원래 데이터 유지
                worksheet = workbook.create_sheet(title=sheet_name)

                # 헤더 추가
                headers = list(df.columns)
                worksheet.append(headers)

                # 한 행 비우기
                worksheet.append([None] * len(headers))

                # 데이터 추가
                for row in dataframe_to_rows(df, index=False, header=False):
                    worksheet.append(row)

        # 엑셀 파일 저장
        workbook.save(output_file_name)
        print(f"[INFO] 모든 시트가 저장되었습니다: {output_file_name}")

    except KeyError as e:
        print(f"[ERROR] 열 '{highlight_column}'이(가) 존재하지 않습니다: {e}")
    except Exception as e:
        print(f"[ERROR] 시트 처리 중 오류 발생: {e}")
        raise





def delete_rows_by_condition(sheet, condition_column, condition_value):
    """
    특정 열의 값이 조건과 일치하는 행을 삭제합니다.

    Parameters:
        sheet (pd.DataFrame): DataFrame 객체 (엑셀 시트 데이터).
        condition_column (str): 조건을 검사할 열 이름.
        condition_value (str): 삭제할 조건 값.

    Returns:
        pd.DataFrame: 조건에 따라 행이 삭제된 DataFrame.
    """
    try:
        # 조건에 맞는 행 확인
        rows_to_delete = sheet[sheet[condition_column] == condition_value]

        # 삭제될 행갯수
        rows_to_delete_count = len(rows_to_delete)
        
        # 로그 출력 (삭제 대상 확인)
        if not rows_to_delete.empty:
            logger.log(f"{condition_value} 으로 삭제할 행 갯수: {rows_to_delete_count}")
            # logger.log(f"삭제 대상:")
            # print(rows_to_delete)
        else:
            logger.log(f"{condition_value} 행이 없어 삭제할 행이 없습니다.")

        # 조건에 맞는 행 삭제
        filtered_sheet = sheet[sheet[condition_column] != condition_value]

        return filtered_sheet, rows_to_delete_count

    except KeyError as e:
        logger.log(f"열 이름 오류: {e}")
        raise
    except Exception as e:
        logger.log(f"행 삭제 중 오류 발생: {e}")
        raise



